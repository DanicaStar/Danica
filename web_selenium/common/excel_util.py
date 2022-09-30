# -*- coding: utf-8 -*-
# @Time    : 2022/9/23 2:08 下午
# @Author  : danica
# @FileName: excel_util.py
# @Software: PyCharm

import os
import openpyxl


class ExcelUtil:
    # 获取当前项目目录的绝对路径
    def get_object_path(self):
        return os.path.abspath(os.path.dirname(__file__)).split('common')[0]

    def read_excel(self):
        excel_path = self.get_object_path() + 'data/input.xlsx'
        # 加载excel工作薄
        wb = openpyxl.load_workbook(excel_path)
        # 获取sheet对象
        sheet = wb['Sheet1']
        # 获取excel的行数和列数
        rows = sheet.max_row
        cols = sheet.max_column
        data_input = []
        for row in range(2, rows + 1):
            data_lst = []
            for col in range(1, cols + 1):
                cell = sheet.cell(row, col).value
                data_lst.append(cell)
            data_input.append(data_lst)
        return data_input


if __name__ == '__main__':
    print(ExcelUtil().read_excel())
