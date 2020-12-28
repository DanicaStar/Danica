# -*- coding:utf-8 -*-
# @Time : 2020-12-26 11:18
# @Author: Danica
# @File : excel_login2.py
import pytest
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

class ExcelData:
    '''
    def test_get_data(self):
        data=[]
        wk=openpyxl.load_workbook('excel_login.xlsx')
        sheet=wk['Sheet2']
        sheet_rows=sheet.iter_rows()
        for item in sheet_rows:     iter_rows返回的是迭代器，没办法从第二行开始去数据，所以会打印出['user', 'pwd']
            list=[]
            for col in item:
                list.append(col.value)
            data.append(list)
        print(data)
    '''
    def get_data(self):
        data=[]
        wk=openpyxl.load_workbook('excel_login.xlsx')
        sheet=wk['Sheet2']
        rows=sheet.max_row
        cols=sheet.max_column
        # for row in range(2,rows+1):
        #     list= {}
        #     for col in range(1,cols+1):
        #         values=sheet.cell(row,col).value
        #         list[sheet.cell(1,col).value]=values
        #         data.append(list)             打印的结果为字典格式
        for row in range(2, rows + 1):
            list =[]
            for col in range(1, cols + 1):
                values = sheet.cell(row, col).value
                list.append(values)
            data.append(list)
        print(data)    #打印结果为一个列表
        return data


class  TestLogin:
    def setup_class(self):
        self.driver=webdriver.Chrome()
        url='http://localhost:8080/jpress/user/login'
        self.driver.maximize_window()
        self.driver.get(url)
    @pytest.mark.parametrize('data',ExcelData().get_data())
    def test_login(self,data):
        user = self.driver.find_element_by_name('user')
        user.clear()
        user.send_keys(data[0])
        passwd = self.driver.find_element_by_name('pwd')
        passwd.clear()
        passwd.send_keys(data[1])
        button = self.driver.find_element_by_class_name('btn')
        button.click()
        WebDriverWait(self.driver, 5).until(EC.alert_is_present())
        alert = self.driver.switch_to.alert
        assert alert.text == data[2]
        alert.accept()

if __name__ == '__main__':
    pytest.main(['-sv','excel_login2.py'])



