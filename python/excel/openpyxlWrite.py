from openpyxl import load_workbook
class ReadExcel:
    def __init__(self,filename):
        self.filename=filename
        self.wk=load_workbook(self.filename)
    def resdExdel(self,sheetname):
        sheet=self.wk[sheetname]
        rows=sheet.max_row
        cols=sheet.max_column
        data=[]
        for x in range(2,rows+1):
            value={}
            for y in range(1,cols+1):
                key=sheet.cell(1,y).value
                value[key]=sheet.cell(x,y).value
            data.append(value)
        return data

if __name__ == '__main__':
    filename= 'test22.xlsx'
    excel=ReadExcel(filename).resdExdel('Sheet1')
    for data in excel:
        print(data)
