'''
http://www.boohee.com/food/
http://www.boohee.com/food/group/1
http://www.boohee.com/food/group/2
http://www.boohee.com/food/group/1?page=2

http://www.boohee.com/food/view_menu
http://www.boohee.com/food/view_menu?page=2
'''

from gevent import monkey
monkey.patch_all()  #让程序变成异步模式
from gevent.queue import Queue
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl import load_workbook
import requests,gevent,csv
work=Queue()
url_1='http://www.boohee.com/food/group/{type}?page={page}'
url_2='http://www.boohee.com/food/view_menu?page={page}'
for i in range(1,3):
    for j in range(1,3):
        real_url=url_1.format(type=i,page=j)
        work.put_nowait(real_url)

for x in range(3):
    real_url=url_2.format(page=x)
    work.put_nowait(real_url)

def boohe_spider():
    headers = {'user-agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.87 Safari/537.36'}
    while not work.empty():    #当队列不为空，执行下面的代码
        url=work.get_nowait()     #使用get_nowait()方法将请求url从队列中取出来
        response=requests.get(url=url,headers=headers)
        print('第{}个食物的第{}页'.format(i,j))
        soup=BeautifulSoup(response.text,'html.parser')
        foods=soup.find_all('li',class_="item clearfix")
        result_list = []
        for food in foods:
            food_name=food.find_all('a')[1]['title']
            food_calorie=food.find('p').text
            food_url='http://www.boohee.com' + food.find_all('a')[1]['href']
            result_list.append([food_name,food_calorie,food_url])


        xlsx(result_list)

# xlsx写入
def xlsx(li):
    try:
        wb = load_workbook("13_boohe.xlsx")
    except:
        wb = Workbook()
    finally:
        try:
            ws = wb[li[0][0]]  # 若工作表不存在就创建一个当前类别命名的工作表
        except Exception as e:
            ws = wb.create_sheet(li[0][0], index=0)
            ws.append([ "食物名称", "热量", "链接"])
        finally:
            for x in li:
                ws.append(x)  # 按行写入数据
    wb.save("13_result.xlsx")
    wb.close()


'''
csv写入
            with open('13_boohe_spider.csv','a') as file:
                writer=csv.writer(file)
                writer.writerow([food_name,food_calorie,food_url])
'''





list_task=[]
for i in range(1,6):
    task=gevent.spawn(boohe_spider)  #使用event.spawn()方法创建执行的boohe_spider()爬虫任务
    list_task.append(task)
gevent.joinall(list_task)   #使用gevent.joinall()方法启动协程任务列表里task_list的所有爬虫任务，让爬虫开始爬取网站信息




