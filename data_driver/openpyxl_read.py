# -*- coding: utf-8 -*-
# @Time    : 2022/9/22 5:35 下午
# @Author  : danica
# @FileName: openpyxl_read.py
# @Software: PyCharm


from openpyxl import load_workbook


class DoExcel:
    def __init__(self, filename):
        """:param filename: excel文件名"""
        self.file = filename
        self.wk = load_workbook(self.file)

    def do_excel(self, sheetname):
        """
        :param sheetname: 工作簿名称
        :return:
        """
        sheet = self.wk[sheetname]
        max_row = sheet.max_row  # 最大行
        max_column = sheet.max_column  # 最大列
        data = []  # 定义一个空列表,用于存储所有数据
        for r in range(2, max_row + 1):
            subdata = {}  # 定义一个字典,用于存储每行数据
            for c in range(1, max_column + 1):
                key = sheet.cell(1, c).value  # 取第一行表头数据
                subdata[key] = sheet.cell(r, c).value  # 字典格式，表头作为key
            data.append(subdata)
        return data


if __name__ == '__main__':
    file_name = "./data/test22.xlsx"
    datas = DoExcel(file_name).do_excel('Sheet1')
    for data in datas:
        print(data)
